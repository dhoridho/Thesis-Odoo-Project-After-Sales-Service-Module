from odoo import _, api, fields, models
from odoo.exceptions import UserError, ValidationError
from datetime import datetime, timedelta, date

import math
import tempfile
import binascii
import openpyxl
import io
import base64

import logging

_logger = logging.getLogger(__name__)

JOB_ESTIMATE_DICT = {
    'External ID': {
        'field': 'external_id',
        'type': 'Char',
        'object': '',
    }, 
    'Project': {
        'field': 'project_id',
        'type': 'Many2one',
        'object': 'project.project',
    }, 
    'Planned Start Date': {
        'field': 'start_date', 
        'type': 'Date',
        'object': '',
    }, 
    'Planned End Date': {
        'field': 'end_date', 
        'type': 'Date',
        'object': '',
    },
}

PROJECT_SCOPE_DICT = {
    'Project Scope': {
        'field': 'project_scope', 
        'type': 'Many2one',
        'object': 'project.scope.line',
    }, 
    'Description': {
        'field': 'description', 
        'type': 'Text',
        'object': '',
    }, 
}

SECTION_DICT = {
    'Project Scope': {
        'field': 'project_scope', 
        'type': 'Many2one',
        'object': 'project.scope.line',
    }, 
    'Section': {
        'field': 'section_name',
        'type': 'Many2one',
        'object': 'section.line',
    }, 
    'Description': {
        'field': 'description', 
        'type': 'Text',
        'object': '',
    }, 
    'Quantity': {
        'field': 'quantity',
        'type': 'Float',
        'object': '',
    }, 
    'Unit of Measure': {
        'field': 'uom_id',
        'type': 'Many2one',
        'object': 'uom.uom',
    }, 
}
			
VARIABLE_DICT = {
    'Project Scope': {
        'field': 'project_scope', 
        'type': 'Many2one',
        'object': 'project.scope.line',
    }, 
    'Section': {
        'field': 'section_name',
        'type': 'Many2one',
        'object': 'section.line',
    }, 
    'Variable': {
        'field': 'variable_name',
        'type': 'Many2one',
        'object': 'variable.template',
    }, 
    'Quantity': {
        'field': 'variable_quantity',
        'type': 'Float',
        'object': '',
    }, 
    'Unit of Measure': {
        'field': 'variable_uom',
        'type': 'Many2one',
        'object': 'uom.uom',
    }, 
}

MATERIAL_ESTIMATION_DICT = {
    'Project Scope': {
        'field': 'project_scope', 
        'type': 'Many2one',
        'object': 'project.scope.line',
    }, 
    'Section': {
        'field': 'section_name',
        'type': 'Many2one',
        'object': 'section.line',
    }, 
    'Group of Product': {
        'field': 'group_of_product',
        'type': 'Many2one',
        'object': 'group.of.product',
    }, 
    'Product': {
        'field': 'product_id',
        'type': 'Many2one',
        'object': 'product.product',
    }, 
    'Description': {
        'field': 'description', 
        'type': 'Text',
        'object': '',
    }, 
    'Coefficient': {
        'field': 'coefficient',
        'type': 'Float',
        'object': '',
    }, 
    'Quantity': {
        'field': 'quantity',
        'type': 'Float',
        'object': '',
    }, 
    'Unit of Measure': {
        'field': 'uom_id',
        'type': 'Many2one',
        'object': 'uom.uom',
    }, 
    'Unit Price': {
        'field': 'unit_price',
        'type': 'Float',
        'object': '',
    }, 
}

LABOUR_ESTIMATION_DICT = {
    'Project Scope': {
        'field': 'project_scope', 
        'type': 'Many2one',
        'object': 'project.scope.line',
    }, 
    'Section': {
        'field': 'section_name',
        'type': 'Many2one',
        'object': 'section.line',
    }, 
    'Group of Product': {
        'field': 'group_of_product',
        'type': 'Many2one',
        'object': 'group.of.product',
    }, 
    'Product': {
        'field': 'product_id',
        'type': 'Many2one',
        'object': 'product.product',
    }, 
    'Description': {
        'field': 'description', 
        'type': 'Text',
        'object': '',
    }, 
    'Contractors': {
        'field': 'contractors',
        'type': 'Integer',
        'object': '',
    }, 
    'Coefficient': {
        'field': 'coefficient',
        'type': 'Float',
        'object': '',
    }, 
    'Time': {
        'field': 'time',
        'type': 'Float',
        'object': '',
    }, 
    # 'Quantity': {
    #     'field': 'quantity',
    #     'type': 'Float',
    #     'object': '',
    # },
    'Unit of Measure': {
        'field': 'uom_id',
        'type': 'Many2one',
        'object': 'uom.uom',
    }, 
    'Unit Price': {
        'field': 'unit_price',
        'type': 'Float',
        'object': '',
    }, 
}

OVERHEAD_ESTIMATION_DICT = {
    'Project Scope': {
        'field': 'project_scope', 
        'type': 'Many2one',
        'object': 'project.scope.line',
    }, 
    'Section': {
        'field': 'section_name',
        'type': 'Many2one',
        'object': 'section.line',
    }, 
    'Overhead Category': {
        'field': 'overhead_catagory',
        'type': 'Selection',
        'object': {
            'Product': 'product',
            'Petty Cash': 'petty cash',
            'Cash Advance': 'cash advance',
            'Fuel': 'fuel',
        },
    }, 
    'Group of Product': {
        'field': 'group_of_product',
        'type': 'Many2one',
        'object': 'group.of.product',
    }, 
    'Product': {
        'field': 'product_id',
        'type': 'Many2one',
        'object': 'product.product',
    }, 
    'Description': {
        'field': 'description', 
        'type': 'Text',
        'object': '',
    }, 
    'Coefficient': {
        'field': 'coefficient',
        'type': 'Float',
        'object': '',
    }, 
    'Quantity': {
        'field': 'quantity',
        'type': 'Float',
        'object': '',
    }, 
    'Unit of Measure': {
        'field': 'uom_id',
        'type': 'Many2one',
        'object': 'uom.uom',
    }, 
    'Unit Price': {
        'field': 'unit_price',
        'type': 'Float',
        'object': '',
    }, 
}

INTERNAL_ASSET_ESTIMATION_DICT = {
    'Project Scope': {
        'field': 'project_scope', 
        'type': 'Many2one',
        'object': 'project.scope.line',
    }, 
    'Section': {
        'field': 'section_name',
        'type': 'Many2one',
        'object': 'section.line',
    }, 
    'Asset Category': {
        'field': 'asset_category_id',
        'type': 'Many2one',
        'object': 'maintenance.equipment.category',
    }, 
    'Asset': {
        'field': 'asset_id',
        'type': 'Many2one',
        'object': 'maintenance.equipment',
    }, 
    'Description': {
        'field': 'description', 
        'type': 'Text',
        'object': '',
    }, 
    'Coefficient': {
        'field': 'coefficient',
        'type': 'Float',
        'object': '',
    }, 
    'Quantity': {
        'field': 'quantity',
        'type': 'Float',
        'object': '',
    }, 
    'Unit of Measure': {
        'field': 'uom_id',
        'type': 'Many2one',
        'object': 'uom.uom',
    }, 
    'Unit Price': {
        'field': 'unit_price',
        'type': 'Float',
        'object': '',
    }, 
}

EQUIPMENT_ESTIMATION_DICT = {
    'Project Scope': {
        'field': 'project_scope', 
        'type': 'Many2one',
        'object': 'project.scope.line',
    }, 
    'Section': {
        'field': 'section_name',
        'type': 'Many2one',
        'object': 'section.line',
    }, 
    'Group of Product': {
        'field': 'group_of_product',
        'type': 'Many2one',
        'object': 'group.of.product',
    }, 
    'Product': {
        'field': 'product_id',
        'type': 'Many2one',
        'object': 'product.product',
    }, 
    'Description': {
        'field': 'description', 
        'type': 'Text',
        'object': '',
    }, 
    'Coefficient': {
        'field': 'coefficient',
        'type': 'Float',
        'object': '',
    }, 
    'Quantity': {
        'field': 'quantity',
        'type': 'Float',
        'object': '',
    }, 
    'Unit of Measure': {
        'field': 'uom_id',
        'type': 'Many2one',
        'object': 'uom.uom',
    }, 
    'Unit Price': {
        'field': 'unit_price',
        'type': 'Float',
        'object': '',
    }, 
}

SUBCON_ESTIMATION_DICT = {
    'Project Scope': {
        'field': 'project_scope', 
        'type': 'Many2one',
        'object': 'project.scope.line',
    }, 
    'Section': {
        'field': 'section_name',
        'type': 'Many2one',
        'object': 'section.line',
    }, 
    'Job SubCon': {
        'field': 'variable', 
        'type': 'Many2one',
        'object': 'variable.template',
    }, 
    'Description': {
        'field': 'description', 
        'type': 'Text',
        'object': '',
    }, 
    'Coefficient': {
        'field': 'coefficient',
        'type': 'Float',
        'object': '',
    }, 
    'Quantity': {
        'field': 'quantity',
        'type': 'Float',
        'object': '',
    }, 
    'Unit of Measure': {
        'field': 'uom_id',
        'type': 'Many2one',
        'object': 'uom.uom',
    }, 
    'Unit Price': {
        'field': 'unit_price',
        'type': 'Float',
        'object': '',
    }, 
}


class ExportJobEstimate(models.TransientModel):
    _name = 'export.job.estimate'
    _description = 'Export Job Estimate'

    type = fields.Selection([
        ('variable', 'BOQ With Variable'),
        ('non_variable', 'BOQ Without Variable'),
        ('all', 'All (Variable and Estimation)'),
    ], required=True, default='variable')
    job_estimate_id = fields.Many2one('job.estimate', string='BOQ')

    def action_export(self):
        if self.job_estimate_id:
            record = self.job_estimate_id
            sheet_dict = {
                'Project Scope': {'dict': PROJECT_SCOPE_DICT,'field': 'project_scope_ids'},
                'Section': {'dict': SECTION_DICT,'field': 'section_ids'},
                'Variable': {'dict': VARIABLE_DICT,'field': 'variable_ids'},
                'Material Estimation': {'dict': MATERIAL_ESTIMATION_DICT,'field': 'material_estimation_ids'},
                'Labour Estimation': {'dict': LABOUR_ESTIMATION_DICT,'field': 'labour_estimation_ids'},
                'Overhead Estimation': {'dict': OVERHEAD_ESTIMATION_DICT,'field': 'overhead_estimation_ids'},
                'Internal Asset Estimation': {'dict': INTERNAL_ASSET_ESTIMATION_DICT,'field': 'internal_asset_ids'},
                'Equipment Estimation': {'dict': EQUIPMENT_ESTIMATION_DICT,'field': 'equipment_estimation_ids'},
                'Subcon Estimation': {'dict': SUBCON_ESTIMATION_DICT,'field': 'subcon_estimation_ids'},
            }
            def prepare_data(obj, dictionary):
                def find_key_by_value(d, search_value):
                    for key, value in d.items():
                        if value == search_value:
                            return key
                    return None 
                data = []
                for k, v in dictionary.items():
                    if v['type'] == 'Many2one':
                        data.append( obj[v['field']].name )
                    elif v['type'] == 'Selection':
                        _key = find_key_by_value(v['object'], obj[v['field']])
                        if _key:
                            data.append(_key)
                        else:
                            data.append('')
                    else :
                        data.append( obj[v['field']] )
                return data

            workbook = openpyxl.Workbook()
            # BOQ
            sheet = workbook.active
            sheet.title = "BOQ"
            headers = [k for k, v in JOB_ESTIMATE_DICT.items()]
            sheet.append(headers)

            job_estimate = prepare_data(record, JOB_ESTIMATE_DICT)

            job_estimate[0] = 'external_id_' + str(record.id)
            record.external_id = 'external_id_' + str(record.id)

            sheet.append(job_estimate)
            for i, x in enumerate(job_estimate):
                sheet.column_dimensions[chr(65+i)].width = 20

            for sheet_name, val in sheet_dict.items():
                if self.type == 'non_variable' and sheet_name == 'Variable': 
                    continue
                if self.type == 'variable' and sheet_name in ['Material Estimation', 'Labour Estimation', 'Overhead Estimation', 'Internal Asset Estimation', 'Equipment Estimation', 'Subcon Estimation'] : 
                    continue

                sheet = workbook.create_sheet(sheet_name)
                headers = [k for k, v in val['dict'].items()]
                if self.type == 'all':
                    if sheet_name in ['Project Scope', 'Section', 'Variable', 'Material Estimation', 'Labour Estimation', 'Overhead Estimation', 'Internal Asset Estimation', 'Equipment Estimation', 'Subcon Estimation']:
                        headers.append('Subtotal')
                    if sheet_name in ['BOQ']:
                        headers.append('Total BOQ')
                sheet.append(headers)

                for i, x in enumerate(headers):
                    sheet.column_dimensions[chr(65+i)].width = 20
                for item in record[val['field']]:
                    data = prepare_data(item, val['dict'])

                    if self.type == 'all':
                        if sheet_name in ['Project Scope', 'Section', 'Variable', 'Material Estimation', 'Labour Estimation', 'Overhead Estimation', 'Internal Asset Estimation', 'Equipment Estimation', 'Subcon Estimation']:
                            data.append( item.subtotal )
                        if sheet_name in ['BOQ']:
                            data.append( item.total_job_estimate )
                    sheet.append(data)

            # Save the workbook to a temporary buffer
            excel_buffer = io.BytesIO()
            workbook.save(excel_buffer)
            excel_buffer.seek(0)

            # Create and associate an attachment
            filename = "boq_exported_data_"+ self.type +".xlsx"
            attachment_data = {
                'name': filename,
                'datas': base64.encodestring(excel_buffer.read()),
                'res_model': 'job.estimate',
                'res_id': self.id,  # Assuming you are exporting a single record
            }
            attachment_id = self.env['ir.attachment'].create(attachment_data)

            # Return a response to download the attachment
            return {
                'type': 'ir.actions.act_url',
                'url': '/web/content/ir.attachment/%s/datas/%s' % (attachment_id.id, filename),
                'target': 'self',
            }


class UploadJobEstimate(models.TransientModel):
    _name = "upload.job.estimate"
    _description = "Upload Job Estimate"

    type = fields.Selection([
        ('variable', 'BOQ With Variable'),
        ('non_variable', 'BOQ Without Variable'),
    ], required=True, default='variable')
    attachment_file = fields.Binary('Attachment File')
    file_name = fields.Char('File Name')
    description = fields.Html(string='Notes')

    def switch_sheet(self, workbook, index):
        workbook.active = index
        sheet_names = workbook.sheetnames
        sheet_title = sheet_names[index]
        sheet = workbook[sheet_title]
        
        keys = [cell.value for cell in sheet[1]]
        rows = [row for row in sheet.iter_rows(min_row=2, values_only=True)]
        
        return keys, rows
    
    def find_data(self, sheet_name, model, model_name, name):
        res = self.env[model].search([('name', '=', name)], limit = 1)
        if not res:
            raise ValidationError(_('[Sheet %s] %s with name %s Not Found!') % (sheet_name, model_name, name))

        return res[0]

    def find_or_create(self, workbook, keys, row):
        update_data = False
        line = dict(zip(keys, row))

        external_id = line.get('External ID', False)
        if external_id:
            res = self.env['job.estimate'].search([('external_id', '=', external_id)], limit = 1)
            if res:
                if res.state_new not in 'draft':
                     raise ValidationError(_('Cannot update the BOQ "%s" with external id "%s" because the state not in draft.') % (res.name, external_id))
                else:
                    res.project_scope_ids.unlink()
                    res.section_ids.unlink()
                    res.variable_ids.unlink()
                    res.material_estimation_ids.unlink()
                    res.labour_estimation_ids.unlink()
                    res.overhead_estimation_ids.unlink()
                    res.internal_asset_ids.unlink()
                    res.equipment_estimation_ids.unlink()
                    res.subcon_estimation_ids.unlink()
                    update_data = True
            else:
                raise ValidationError(_('Cannot find BOQ with provided External ID.'))
        else:
            res = self.create_job_estimate(workbook, keys, row)
            res.external_id = external_id
        

        return res, update_data

    def create_job_estimate(self, workbook, keys, row):
        data = self.prepare_data_lines( workbook, JOB_ESTIMATE_DICT, keys, row )
        res = self.env['job.estimate'].create(data)
        return res
        
    def prepare_data_lines(self, workbook, dictionary, keys, row):
        line = dict(zip(keys, row))
        data = {}

        for key, val in dictionary.items():
            if line.get(key) == None:
                return {}
            if key == 'External ID':
                continue
            if val['type'] == 'Selection':
                data[val['field']] = val['object'][line.get(key)]
            # elif val['type'] == 'Date':
            #     cell_value = line.get(key)
            #     if isinstance(cell_value, float):
            #         days_since_epoch = int(cell_value)
            #         base_date = datetime(1899, 12, 30)
            #         date_value = base_date + timedelta(days=days_since_epoch)
            #         data[val['field']] = date_value
            elif val['type'] == 'Many2one':
                active_sheet = workbook.active
                obj = self.find_data(active_sheet.title, val['object'], key, line.get(key))
                data[val['field']] = obj.id
            else:
                data[val['field']] = line.get(key)

        return data

    def load_workbook(self):
        import_name_extension = self.file_name.split('.')[1]
        if import_name_extension not in ['xls', 'xlsx']:
            raise ValidationError('The upload file is using the wrong format. Please upload your file in xlsx or xls format.')
        
        fp = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
        fp.write(binascii.a2b_base64(self.attachment_file))
        fp.seek(0)
        
        workbook = openpyxl.load_workbook(fp.name, data_only=True)
        return workbook
        
    def import_action(self):
        workbook = self.load_workbook()
        job_estimate_dict = {
            'project_scope' : {}
        }
        def check_project_scope_section(project_scope_id, section_id):
            project_scope = self.env['project.scope.line'].search([('id', '=', project_scope_id)], limit = 1)
            section = self.env['section.line'].search([('id', '=', section_id)], limit = 1)
            if project_scope and section:
                # check Project Scope
                if job_estimate_dict['project_scope'].get(project_scope_id, False):
                    # check Section
                    if not job_estimate_dict['project_scope'][project_scope_id]['section'].get(section_id, False):
                        raise ValidationError(_('[Sheet %s] Section %s Not Listed in Section Tab for Project Scope %s!') % (workbook.active.title, section.name, project_scope.name))
                else:
                    raise ValidationError(_('[Sheet %s] Project Scope Not Listed in Project Scope Tab!') % (workbook.active.title))

        def get_setion(project_scope_id):
            res = {}
            section_ids = self.env['section.line'].search([('project_scope', '=', project_scope_id)])
            for section_id in section_ids:
                res[section_id.id] = {}
            return res

        def check_boq_type(type, sheet_names):
            if type == 'variable' and not ('Variable' in sheet_names) or \
                type == 'non_variable' and ('Variable' in sheet_names):
                raise ValidationError(_('BOQ type and file missmatch!'))

        check_boq_type(self.type, workbook.sheetnames)
        # job_estimate = {}
        keys, rows = self.switch_sheet(workbook, 0)
        # for row in rows:

        job_estimate, update_data = self.find_or_create(workbook, keys, rows[0])

        # project_scope
        keys, rows = self.switch_sheet(workbook, 1)
        for row in rows:
            data = self.prepare_data_lines(workbook, PROJECT_SCOPE_DICT, keys, row)
            if not data : continue
            project_scope_id = data[ PROJECT_SCOPE_DICT['Project Scope']['field'] ]

            job_estimate_dict['project_scope'][project_scope_id] = { 'section': {} }
            job_estimate.write({
                'project_scope_ids':[(0, 0, data)]
            })
        
        # section
        keys, rows = self.switch_sheet(workbook, 2)
        for row in rows:
            data = self.prepare_data_lines(workbook, SECTION_DICT, keys, row)
            if not data : continue
            project_scope_id = data[ SECTION_DICT['Project Scope']['field'] ]
            section_id = data[ SECTION_DICT['Section']['field'] ] 

            # check Project Scope
            if job_estimate_dict['project_scope'].get(project_scope_id, False):
                job_estimate_dict['project_scope'][ project_scope_id ]['section'][ section_id ] = {'id': section_id}
            else:
                raise ValidationError(_('[Sheet %s] Project Scope Not Listed in Project Scope Tab!') % (workbook.active.title))

            job_estimate.write({
                'section_ids':[(0, 0, data)]
            })
        
        # if with variable
        # variable
        if self.type == 'variable':
            keys, rows = self.switch_sheet(workbook, 3)
            for row in rows:
                data = self.prepare_data_lines(workbook, VARIABLE_DICT, keys, row)
                if not data : continue
                project_scope_id = data[ VARIABLE_DICT['Project Scope']['field'] ]
                section_id = data[ VARIABLE_DICT['Section']['field'] ] 
                # check
                check_project_scope_section(project_scope_id, section_id)

                job_estimate.write({
                    'variable_ids': [(0, 0, data)]
                })
                job_estimate.update_material()
        
        # if without variable
        # non_variable
        elif self.type == 'non_variable':
            
            keys, rows = self.switch_sheet(workbook, 3)
            for row in rows:
                data = self.prepare_data_lines(workbook, MATERIAL_ESTIMATION_DICT, keys, row)
                if not data : continue
                project_scope_id = data[ MATERIAL_ESTIMATION_DICT['Project Scope']['field'] ]
                section_id = data[ MATERIAL_ESTIMATION_DICT['Section']['field'] ] 
                # check
                check_project_scope_section(project_scope_id, section_id)

                job_estimate.write({
                    'material_estimation_ids': [(0, 0, data)]
                })

            keys, rows = self.switch_sheet(workbook, 4)
            for row in rows:
                data = self.prepare_data_lines(workbook, LABOUR_ESTIMATION_DICT, keys, row)
                if not data : continue
                project_scope_id = data[ LABOUR_ESTIMATION_DICT['Project Scope']['field'] ]
                section_id = data[ LABOUR_ESTIMATION_DICT['Section']['field'] ] 
                # check
                check_project_scope_section(project_scope_id, section_id)

                job_estimate.write({
                    'labour_estimation_ids': [(0, 0, data)]
                })

            keys, rows = self.switch_sheet(workbook, 5)
            for row in rows:
                data = self.prepare_data_lines(workbook, OVERHEAD_ESTIMATION_DICT, keys, row)
                if not data : continue
                project_scope_id = data[ OVERHEAD_ESTIMATION_DICT['Project Scope']['field'] ]
                section_id = data[ OVERHEAD_ESTIMATION_DICT['Section']['field'] ] 
                # check
                check_project_scope_section(project_scope_id, section_id)

                job_estimate.write({
                    'overhead_estimation_ids': [(0, 0, data)]
                })

            keys, rows = self.switch_sheet(workbook, 6)
            for row in rows:
                data = self.prepare_data_lines(workbook, INTERNAL_ASSET_ESTIMATION_DICT, keys, row)
                if not data : continue
                project_scope_id = data[ INTERNAL_ASSET_ESTIMATION_DICT['Project Scope']['field'] ]
                section_id = data[ INTERNAL_ASSET_ESTIMATION_DICT['Section']['field'] ] 
                # check
                check_project_scope_section(project_scope_id, section_id)

                job_estimate.write({
                    'internal_asset_ids': [(0, 0, data)]
                })

            keys, rows = self.switch_sheet(workbook, 7)
            for row in rows:
                data = self.prepare_data_lines(workbook, EQUIPMENT_ESTIMATION_DICT, keys, row)
                if not data : continue
                project_scope_id = data[ EQUIPMENT_ESTIMATION_DICT['Project Scope']['field'] ]
                section_id = data[ EQUIPMENT_ESTIMATION_DICT['Section']['field'] ] 
                # check
                check_project_scope_section(project_scope_id, section_id)

                job_estimate.write({
                    'equipment_estimation_ids': [(0, 0, data)]
                })

            keys, rows = self.switch_sheet(workbook, 8)
            for row in rows:
                data = self.prepare_data_lines(workbook, SUBCON_ESTIMATION_DICT, keys, row)
                if not data : continue
                project_scope_id = data[ SUBCON_ESTIMATION_DICT['Project Scope']['field'] ]
                section_id = data[ SUBCON_ESTIMATION_DICT['Section']['field'] ] 
                # check
                check_project_scope_section(project_scope_id, section_id)

                job_estimate.write({
                    'subcon_estimation_ids': [(0, 0, data)]
                })
        
        if not update_data:
            job_estimate.onchange_project_id_when_import()
        else:
            job_estimate.exist_main_contract()

        job_estimate.material_estimation_ids._onchange_coefficient()
        job_estimate.material_estimation_ids.onchange_quantity()

        job_estimate.labour_estimation_ids._onchange_coefficient()
        job_estimate.labour_estimation_ids.onchange_quantity()

        job_estimate.overhead_estimation_ids._onchange_coefficient()
        job_estimate.overhead_estimation_ids.onchange_quantity()

        job_estimate.internal_asset_ids._onchange_coefficient()
        job_estimate.internal_asset_ids.onchange_quantity()

        job_estimate.equipment_estimation_ids._onchange_coefficient()
        job_estimate.equipment_estimation_ids.onchange_quantity()

        job_estimate.subcon_estimation_ids._onchange_coefficient()
        job_estimate.subcon_estimation_ids.onchange_quantity()

        job_estimate._compute_approving_customer_matrix()
        job_estimate.onchange_approving_matrix_lines()

        if job_estimate:
            return {
                    'name': _('BOQ'),
                    'type': 'ir.actions.act_window',
                    'res_model': 'job.estimate',
                    'view_id': self.env.ref('equip3_construction_sales_operation.project_scope_estimation_inherit').id,
                    'view_type': 'form',
                    'view_mode': 'form',
                    'res_id': job_estimate.id,
                }
        