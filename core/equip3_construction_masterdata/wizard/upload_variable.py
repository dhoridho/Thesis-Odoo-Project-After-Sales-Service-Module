from odoo import _, api, fields, models
from odoo.exceptions import UserError, ValidationError
from datetime import datetime, timedelta, date

import math
import tempfile
import binascii
import openpyxl
import io
import base64

VARIABLE_DICT = {
    'External ID': {
        'field': 'external_id',
        'type': 'Char',
        'object': '',
    }, 
    'Name': {
        'field': 'name',
        'type': 'Char',
        'object': '',
    }, 
    'Unit of Measure': {
        'field': 'variable_uom',
        'type': 'Many2one',
        'object': 'uom.uom',
    }, 
    'Branch': {
        'field': 'branch_id', 
        'type': 'Many2one',
        'object': 'res.branch',
    }, 
}

MATERIAL_VARIABLE_DICT = {
    'Group of Product': {
        'field': 'group_of_product',
        'type': 'Many2one',
        'object': 'group.of.product',
    }, 
    'Product': {
        'field': 'product_id',
        'type': 'Many2one',
        'object': 'product.product',
    }, 
    'Description': {
        'field': 'description', 
        'type': 'Text',
        'object': '',
    }, 
    'Quantity': {
        'field': 'quantity',
        'type': 'Float',
        'object': '',
    }, 
    'Unit of Measure': {
        'field': 'uom_id',
        'type': 'Many2one',
        'object': 'uom.uom',
    }, 
    'Unit Price': {
        'field': 'unit_price',
        'type': 'Float',
        'object': '',
    }, 
}

LABOUR_VARIABLE_DICT = {
    'Group of Product': {
        'field': 'group_of_product',
        'type': 'Many2one',
        'object': 'group.of.product',
    }, 
    'Product': {
        'field': 'product_id',
        'type': 'Many2one',
        'object': 'product.product',
    }, 
    'Description': {
        'field': 'description', 
        'type': 'Text',
        'object': '',
    }, 
    'Contractors': {
        'field': 'contractors',
        'type': 'Integer',
        'object': '',
    }, 
    'Time': {
        'field': 'time',
        'type': 'Float',
        'object': '',
    }, 
    'Quantity': {
        'field': 'quantity',
        'type': 'Float',
        'object': '',
    }, 
    'Unit of Measure': {
        'field': 'uom_id',
        'type': 'Many2one',
        'object': 'uom.uom',
    }, 
    'Unit Price': {
        'field': 'unit_price',
        'type': 'Float',
        'object': '',
    }, 
}

OVERHEAD_VARIABLE_DICT = {
    'Overhead Category': {
        'field': 'overhead_catagory',
        'type': 'Selection',
        'object': {
            'Product': 'product',
            'Petty Cash': 'petty_cash',
            'Cash Advance': 'cash_advance',
            'Fuel': 'fuel',
        },
    }, 
    'Group of Product': {
        'field': 'group_of_product',
        'type': 'Many2one',
        'object': 'group.of.product',
    }, 
    'Product': {
        'field': 'product_id',
        'type': 'Many2one',
        'object': 'product.product',
    }, 
    'Description': {
        'field': 'description', 
        'type': 'Text',
        'object': '',
    }, 
    'Quantity': {
        'field': 'quantity',
        'type': 'Float',
        'object': '',
    }, 
    'Unit of Measure': {
        'field': 'uom_id',
        'type': 'Many2one',
        'object': 'uom.uom',
    }, 
    'Unit Price': {
        'field': 'unit_price',
        'type': 'Float',
        'object': '',
    }, 
}

INTERNAL_ASSET_VARIABLE_DICT = {
    'Asset Category': {
        'field': 'asset_category_id',
        'type': 'Many2one',
        'object': 'maintenance.equipment.category',
    }, 
    'Asset': {
        'field': 'asset_id',
        'type': 'Many2one',
        'object': 'maintenance.equipment',
    }, 
    'Description': {
        'field': 'description', 
        'type': 'Text',
        'object': '',
    }, 
    'Quantity': {
        'field': 'quantity',
        'type': 'Float',
        'object': '',
    }, 
    'Unit of Measure': {
        'field': 'uom_id',
        'type': 'Many2one',
        'object': 'uom.uom',
    }, 
    'Unit Price': {
        'field': 'unit_price',
        'type': 'Float',
        'object': '',
    }, 
}

EQUIPMENT_VARIABLE_DICT = {
    'Group of Product': {
        'field': 'group_of_product',
        'type': 'Many2one',
        'object': 'group.of.product',
    }, 
    'Product': {
        'field': 'product_id',
        'type': 'Many2one',
        'object': 'product.product',
    }, 
    'Description': {
        'field': 'description', 
        'type': 'Text',
        'object': '',
    }, 
    'Quantity': {
        'field': 'quantity',
        'type': 'Float',
        'object': '',
    }, 
    'Unit of Measure': {
        'field': 'uom_id',
        'type': 'Many2one',
        'object': 'uom.uom',
    }, 
    'Unit Price': {
        'field': 'unit_price',
        'type': 'Float',
        'object': '',
    }, 
}

SERVICE_VARIABLE_DICT = {
    'Group of Product': {
        'field': 'group_of_product',
        'type': 'Many2one',
        'object': 'group.of.product',
    }, 
    'Product': {
        'field': 'product_id',
        'type': 'Many2one',
        'object': 'product.product',
    }, 
    'Description': {
        'field': 'description', 
        'type': 'Text',
        'object': '',
    }, 
    'Quantity': {
        'field': 'quantity',
        'type': 'Float',
        'object': '',
    }, 
    'Unit of Measure': {
        'field': 'uom_id',
        'type': 'Many2one',
        'object': 'uom.uom',
    }, 
    'Unit Price': {
        'field': 'unit_price',
        'type': 'Float',
        'object': '',
    },  
}

SUBCON_VARIABLE_DICT = {
    'Job SubCon': {
        'field': 'variable', 
        'type': 'Many2one',
        'object': 'variable.template',
    }, 
    'Description': {
        'field': 'description', 
        'type': 'Text',
        'object': '',
    }, 
    'Quantity': {
        'field': 'quantity',
        'type': 'Float',
        'object': '',
    }, 
    'Unit of Measure': {
        'field': 'uom_id',
        'type': 'Many2one',
        'object': 'uom.uom',
    }, 
    'Unit Price': {
        'field': 'unit_price',
        'type': 'Float',
        'object': '',
    }, 
}


class UploadVariable(models.TransientModel):
    _name = "upload.variable.cons"
    _description = "Upload Variable"

    type = fields.Selection([
        ('variable', 'Variable'),
        ('job_subcon', 'Job Subcon')], string='Type')
    
    attachment_file = fields.Binary('Attachment File')
    file_name = fields.Char('File Name')
    description_variable = fields.Html(string="Notes")
    description_subcon = fields.Html(string="Notes")


    def switch_sheet(self, workbook, index):
        workbook.active = index
        sheet_names = workbook.sheetnames
        sheet_title = sheet_names[index]
        sheet = workbook[sheet_title]
        
        keys = [cell.value for cell in sheet[1]]
        rows = [row for row in sheet.iter_rows(min_row=2, values_only=True)]
        
        return keys, rows
    
    def find_data(self, sheet_name, model, model_name, name):
        res = self.env[model].search([('name', '=', name)], limit = 1)
        if not res:
            raise ValidationError(_('[Sheet %s] %s with name %s Not Found!') % (sheet_name, model_name, name))

        return res[0]

    def find_or_create(self, workbook, keys, row):
        update_data = False
        line = dict(zip(keys, row))

        external_id = line.get('External ID', False)
        if external_id:
            res = self.env['variable.template'].search([('external_id', '=', external_id)], limit = 1)
            if res:
                res.material_variable_ids.unlink()
                res.labour_variable_ids.unlink()
                res.overhead_variable_ids.unlink()
                res.asset_variable_ids.unlink()
                res.equipment_variable_ids.unlink()
                res.service_variable_ids.unlink()
                res.subcon_variable_ids.unlink()
                update_data = True
            else:
                res = self.create_variable(workbook, keys, row)
                res.external_id = external_id
        else:
            raise ValidationError(_('External ID must be provided!'))

        return res, update_data

    def create_variable(self, workbook, keys, row):
        data = self.prepare_data_lines( workbook, VARIABLE_DICT, keys, row )
        res = self.env['variable.template'].create(data)
        return res
        
    def prepare_data_lines(self, workbook, dictionary, keys, row):
        line = dict(zip(keys, row))
        data = {}

        for key, val in dictionary.items():
            if line.get(key) == None:
                return {}
            if key == 'External ID':
                continue
            if val['type'] == 'Selection':
                data[val['field']] = val['object'][line.get(key)]
            elif val['type'] == 'Many2one':
                active_sheet = workbook.active
                obj = self.find_data(active_sheet.title, val['object'], key, line.get(key))
                data[val['field']] = obj.id
            else:
                data[val['field']] = line.get(key)

        return data


    def load_workbook(self):
        import_name_extension = self.file_name.split('.')[1]
        if import_name_extension not in ['xls', 'xlsx']:
            raise ValidationError('The upload file is using the wrong format. Please upload your file in xlsx or xls format.')
        
        fp = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
        fp.write(binascii.a2b_base64(self.attachment_file))
        fp.seek(0)
        
        workbook = openpyxl.load_workbook(fp.name, data_only=True)
        return workbook
        
    def import_action(self):
        workbook = self.load_workbook()
        keys, rows = self.switch_sheet(workbook, 0)
        variable, update_data = self.find_or_create(workbook, keys, rows[0])
        
        variable_items = [
            {'sheet_index': 1,'dict': MATERIAL_VARIABLE_DICT,'field': 'material_variable_ids'},
            {'sheet_index': 2,'dict': LABOUR_VARIABLE_DICT,'field': 'labour_variable_ids'},
            {'sheet_index': 4,'dict': OVERHEAD_VARIABLE_DICT,'field': 'overhead_variable_ids'},
            {'sheet_index': 5,'dict': INTERNAL_ASSET_VARIABLE_DICT,'field': 'asset_variable_ids'},
            {'sheet_index': 6,'dict': EQUIPMENT_VARIABLE_DICT,'field': 'equipment_variable_ids'},
            {'sheet_index': 7,'dict': SUBCON_VARIABLE_DICT,'field': 'subcon_variable_ids'},
            ]
        job_subcon_items = [
            {'sheet_index': 1,'dict': MATERIAL_VARIABLE_DICT,'field': 'material_variable_ids'},
            {'sheet_index': 2,'dict': LABOUR_VARIABLE_DICT,'field': 'labour_variable_ids'},
            {'sheet_index': 3,'dict': SERVICE_VARIABLE_DICT,'field': 'service_variable_ids'},
            {'sheet_index': 4,'dict': OVERHEAD_VARIABLE_DICT,'field': 'overhead_variable_ids'},
            {'sheet_index': 6,'dict': EQUIPMENT_VARIABLE_DICT,'field': 'equipment_variable_ids'},
            ]
        
        type_dict = { 'variable': variable_items, 'job_subcon': job_subcon_items}

        if self.type :
            for item in type_dict[ self.type ]:
                keys, rows = self.switch_sheet(workbook, item['sheet_index'])
                for row in rows:
                    data = self.prepare_data_lines(workbook, item['dict'], keys, row)
                    if not data : continue
                    variable.write({
                        item['field']: [(0, 0, data)]
                    })

        variable._calculate_total()
        variable.material_variable_ids.onchange_quantity()
        variable.labour_variable_ids.onchange_quantity()
        variable.overhead_variable_ids.onchange_quantity()
        variable.asset_variable_ids.onchange_quantity()
        variable.equipment_variable_ids.onchange_quantity()
        variable.service_variable_ids.onchange_quantity()
        variable.subcon_variable_ids.onchange_quantity()
        