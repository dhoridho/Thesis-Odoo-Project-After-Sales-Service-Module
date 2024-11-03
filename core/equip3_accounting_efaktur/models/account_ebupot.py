# -*- coding: utf-8 -*-

import time
from odoo import models, fields, api, _
from odoo.exceptions import UserError, ValidationError
from odoo.osv import expression
from odoo.tools.misc import format_date
import base64
import inspect
import os
from datetime import datetime
from dateutil import relativedelta

import openpyxl
from openpyxl import load_workbook


class AccountEBupot(models.Model):
    _name = "account.ebupot"
    _description = "Account E-Bupot"

    year = fields.Many2one('sh.fiscal.year', string='Fiscal Year')
    name = fields.Char(string='eBupot Number', size=15, help="E-Faktur Format xxx-xx-xxxxxxxx")
    is_used = fields.Boolean(string='Is Used',default=False)
    invoice_ids = fields.One2many('account.move','ebupot_id', string='Invoice', domain="[('ebupot_id', 'in', [False,'']), ('state', 'not in', ['cancel','draft']),('move_type', 'in', ['out_invoice','out_refund','out_receipt'])]")
    tax_report_id = fields.Selection(selection=[
                  ('out', 'e-BupotUnifikasi')
                  ], string='Tax Report')
    date_from = fields.Date(string='Start Date')
    date_to = fields.Date(string='End Date')
    partner_id = fields.Many2one('res.partner', string='Vendor')
    company_id = fields.Many2one('res.company', string='Company', default=lambda self: self.env.company.id,readonly=True)
    ebupot_template = fields.Binary('Template', compute="_get_template")
    tahun_pajak = fields.Char(string='Tahun Pajak', size=4)

    _sql_constraints = [
        ('name_uniq', 'unique (name)', 'The Nomor Seri Faktur Pajak is already generated. Please check your Nomor Seri Faktur Pajak')
    ]

    @api.onchange('tahun_pajak')
    def onchange_tahun_pajak(self):
        if self.tahun_pajak and not self.tahun_pajak.isnumeric():
            raise UserError(_("Only numeric inputs allowed!"))

    def _get_template(self):
        dir_name = os.path.dirname(os.path.abspath(inspect.getfile(inspect.currentframe())))
        static_path = os.path.dirname(dir_name)
        path = os.path.join(static_path, 'static/src/file/FORMAT_UPLOAD_EBUPOT_UNIFIKASI_1.xlsx')
        self.ebupot_template = base64.b64encode(open(path, "rb").read())

    def confirm_ebupot(self):
        if self.tahun_pajak and not self.tahun_pajak.isnumeric():
            raise UserError(_("Only numeric inputs allowed!"))
        dir_name = os.path.dirname(os.path.abspath(inspect.getfile(inspect.currentframe())))
        static_path = os.path.dirname(dir_name)
        path = os.path.join(static_path, 'static/src/file/FORMAT_UPLOAD_EBUPOT_UNIFIKASI.xlsx')
        if self.partner_id:
            move_ids = self.env["account.move"].search([
                                                        ('invoice_date', '>=', self.date_from),
                                                        ('invoice_date', '<=', self.date_to),
                                                        ('move_type', '=', 'in_invoice'),
                                                        ('partner_id', '=', self.partner_id.id),
                                                        ('company_id', '=', self.company_id.id)])
        else:
            move_ids = self.env["account.move"].search([
                ('invoice_date', '>=', self.date_from),
                ('invoice_date', '<=', self.date_to),
                ('move_type', '=', 'in_invoice'),
                ('company_id', '=', self.company_id.id)])
        workbook = load_workbook(filename=path)
        nr_line = 0
        tax_line = 0
        for sh in workbook.sheetnames:
            sheet = workbook[sh]
            if sh == "NR":
                no = 3
                pph_taxes_ids = self.env['account.tax'].search([('is_pph','=',True),('pph_type','=','PPH26')]).ids
                for move_id in move_ids:
                    for line_id in move_id.invoice_line_ids:
                        if bool(set(line_id.tax_ids.ids) & set(pph_taxes_ids)):
                            list_kode_objek_pajak = []
                            for kode in move_id.kode_objek_pajak_id:
                                list_kode_objek_pajak.append(kode.name)

                            npwp_penandatanganan = nik_penandatanganan = '-'
                            if line_id.move_id.company_id.is_use_nik_for_ebupot:
                                npwp_nik = 'NIK'
                                nik_penandatanganan = line_id.move_id.company_id.tax_cutter_nik
                            else:
                                npwp_nik = 'NPWP'
                                npwp_penandatanganan = line_id.move_id.company_id.tax_cutter_npwp

                            sheet.insert_rows(no, amount=1)
                            sheet['B'+str(no)] = move_id.invoice_date or "-"
                            sheet['C'+str(no)] = move_id.partner_id.tin or "-"
                            sheet['D'+str(no)] = move_id.partner_id.name or "-"
                            sheet['E'+str(no)] = move_id.partner_id.date or "-"
                            sheet['F'+str(no)] = move_id.partner_id.city or "-"
                            sheet['G'+str(no)] = move_id.partner_id.country_id.name or "-"
                            sheet['H'+str(no)] = move_id.partner_id.no_paspor or "-"
                            sheet['I'+str(no)] = move_id.partner_id.no_kitas or "-"
                            sheet['J'+str(no)] = move_id.partner_id.country_id.code or "-"
                            sheet['K'+str(no)] = ', '.join(list_kode_objek_pajak) or "-"
                            sheet['L'+str(no)] = move_id.company_id.tax_cutter_name.name or "-"
                            sheet['M'+str(no)] = npwp_nik
                            sheet['N'+str(no)] = npwp_penandatanganan
                            sheet['O'+str(no)] = nik_penandatanganan
                            sheet['P'+str(no)] = move_id.partner_id.nin or "-"
                            sheet['Q'+str(no)] = line_id.quantity * line_id.price_unit
                            sheet['R'+str(no)] = "100"
                            sheet['S'+str(no)] = move_id.partner_id.tax_facility or "-"
                            sheet['T'+str(no)] = move_id.partner_id.doc_number or "-"
                            sheet['U'+str(no)] = "-"
                            sheet['V'+str(no)] = move_id.partner_id.doc_number or "-"
                            sheet['W'+str(no)] = move_id.partner_id.other_tax or "-"
                            sheet['X'+str(no)] = move_id.partner_id.desc_other_tax or "-"
                            sheet['Y'+str(no)] = move_id.partner_id.overpayment_processed_by or "-"
                            no+=1
                nr_line = no - 3
            
            if sh == "42152223":
                no = 3
                pph_taxes_ids = self.env['account.tax'].search([('is_pph','=',True),('pph_type','in',['PPH23','PPH4-2','PPH22','PPH15'])]).ids
                for move_id in move_ids:
                    for line_id in move_id.invoice_line_ids:
                        if bool(set(line_id.tax_ids.ids) & set(pph_taxes_ids)):
                            list_kode_objek_pajak = []
                            for kode in move_id.kode_objek_pajak_id:
                                list_kode_objek_pajak.append(kode.name)

                            npwp_penandatanganan = nik_penandatanganan = '-'
                            if line_id.move_id.company_id.is_use_nik_for_ebupot:
                                npwp_nik = 'NIK'
                                nik_penandatanganan = line_id.move_id.company_id.tax_cutter_nik
                            else:
                                npwp_nik = 'NPWP'
                                npwp_penandatanganan = line_id.move_id.company_id.tax_cutter_npwp

                            sheet.insert_rows(no, amount=1)
                            sheet['B'+str(no)] = move_id.invoice_date or "-"
                            sheet['C'+str(no)] = "NPWP"
                            sheet['D'+str(no)] = move_id.partner_id.vat or "-"
                            sheet['E'+str(no)] = "-"
                            sheet['F'+str(no)] = move_id.partner_id.name or "-"
                            sheet['G'+str(no)] = "-"
                            sheet['H'+str(no)] = move_id.partner_id.phone or "-"
                            sheet['I'+str(no)] = ', '.join(list_kode_objek_pajak) or "-"
                            sheet['J'+str(no)] = move_id.company_id.tax_cutter_name.name or "-"
                            sheet['K'+str(no)] = npwp_nik
                            sheet['L'+str(no)] = npwp_penandatanganan
                            sheet['M'+str(no)] = nik_penandatanganan
                            sheet['N'+str(no)] = move_id.partner_id.nin or "-"
                            sheet['O'+str(no)] = line_id.quantity * line_id.price_unit
                            sheet['P'+str(no)] = move_id.partner_id.tax_facility or "-"
                            sheet['Q'+str(no)] = move_id.partner_id.doc_number or "-"
                            sheet['R'+str(no)] = move_id.partner_id.doc_number or "-"
                            sheet['S'+str(no)] = move_id.partner_id.doc_number or "-"
                            sheet['T'+str(no)] = move_id.partner_id.other_tax or "-"
                            sheet['U'+str(no)] = move_id.partner_id.desc_other_tax or "-"
                            sheet['V'+str(no)] = move_id.partner_id.overpayment_processed_by or "-"
                            no+=1
                tax_line = no - 3
                # sheet.row_dimensions[5].height = 15

            if sh == "Dasar Pemotongan":
                no = 3
                pph_taxes_ids = self.env['account.tax'].search([('is_pph','=',True)]).ids
                # sheet.unmerge_cells('B4:E8')
                for move_id in move_ids:
                    for line_id in move_id.invoice_line_ids:
                        if bool(set(line_id.tax_ids.ids) & set(pph_taxes_ids)):
                            jenis_dokumen = '-'
                            if line_id.move_id.kode_dokumen == '1':
                                jenis_dokumen = 'Faktur Pajak'
                            elif line_id.move_id.kode_dokumen == '2':
                                jenis_dokumen = 'Invoice'

                            sheet.insert_rows(no, amount=1)
                            sheet['A'+str(no)] = "-"
                            sheet['B'+str(no)] = "-"
                            sheet['C'+str(no)] = jenis_dokumen
                            sheet['D'+str(no)] = line_id.move_id.kode_dokumen or "-"
                            sheet['E'+str(no)] = move_id.invoice_date or "-"
                            no+=1
                # sheet.merge_cells('B' + str(4+no-3) + ':E' + str(8+no-3))

        masa_pajak = "-"
        if self.date_from and self.date_to:
            r = relativedelta.relativedelta(self.date_to, self.date_from)
            masa_pajak = r.months + (12*r.years)
            #print(masa_pajak)
            
        sheet = workbook["Rekap"]
        sheet['D2'] = self.tahun_pajak or "-"
        sheet['G2'] = masa_pajak
        sheet['G3'] = tax_line
        sheet['G4'] = nr_line

        path1 = os.path.join(static_path, 'static/src/file/FORMAT_UPLOAD_EBUPOT_UNIFIKASI_1.xlsx')
        workbook.save(filename=path1)

        return {
                'type': 'ir.actions.act_url',
                'name': 'ebupot',
                'url': '/web/content/account.ebupot/%s/ebupot_template/FORMAT_UPLOAD_EBUPOT_UNIFIKASI_1.xlsx?download=true' %(self.id),
                'target':'new'
                }


class AccountEBupotGenerate(models.Model):
    _name = "account.ebupot.generate"
    _description = "Account E-Bupot Generate"

    year = fields.Many2one('sh.fiscal.year', string='Fiscal Year')
    start = fields.Char(string='Start', size=11, help="E-Faktur Format xxx-xxxxxxx")
    end = fields.Char(string='End', size=11, help="E-Faktur Format xxx-xxxxxxx")

    # @api.onchange('year')
    # def onchange_year(self):
    #     warning = {}
    #     if self.year and (self.year.isdigit() == False or (self.year.isdigit() == True and len(str(self.year)) > 4)):
    #         self.year = ''
    #         warning = {'title': 'Value Error', 'message': "Invalid year!"}
    #     return {'warning': warning}
        
    def _prepare_ebupot_line(self,prefix,year,start,end):
        list_line=[]
        for x in range(int(start),int(end)+1):
            lines_dict={
                        'year': year,
                        'name': str(self._addnumb(x)),                            
                        }
            list_line.append(lines_dict)
        return list_line

    def _addnumb(self,number):
        numb = str(number)
        while len(numb) < 7:
            numb = str('0') + str(numb)
        return numb

    def confirm(self):
        prefix=''
        years=''
        start_number = ''
        end_number = ''
        for vals in self:
            if vals.start and vals.end:
                if len(vals.start) < 7 or len(vals.end) < 7:
                    raise UserError(_("Please check your code at your Nomor E-Ebupot !"))
            else:
                raise UserError(_("Please check your code at your Nomor E-Ebupot !"))
            start = vals.start
            end = vals.end
            # if start[3] != '-' or end[3] != '-' or start[:3] != end[:3]:
            #     raise UserError(_("Please check your region code and year code at your Nomor E-Ebupot !"))
            prefix = vals.start[:4]
            years = vals.year.id
            if start.isnumeric() == False or end.isnumeric() == False:
                raise UserError(_("Please check your code at your Nomor E-Ebupot !"))
        list_line = self._prepare_ebupot_line(prefix, years, start, end)
        ebupot = self.env['account.ebupot'].create(list_line)
        return {
                'name':  _('E-Bupot'),
                'type': 'ir.actions.act_window',
                'view_mode': 'tree,kanban,form',
                'res_model': 'account.ebupot',
                'views_id': self.env.ref('equip3_accounting_efaktur.view_account_ebupot_tree').id,
                'domain': [('name', 'in', [x['name'] for x in list_line])]
            }
    
class KodeObjekPajak(models.Model):
    _name = "kode.objek.pajak"
    _description = "Kode Objek Pajak"

    name = fields.Char(string='Kode Objek Pajak')
    description = fields.Char(string='Nama Objek Pajak')
    pph_pasal = fields.Char(string='PPH Pasal')
