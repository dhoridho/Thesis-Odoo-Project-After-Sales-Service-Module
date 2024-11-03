# # -*- coding: utf-8 -*-

import os
import tempfile
import base64
from openpyxl import Workbook
from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment, NamedStyle

from dateutil.relativedelta import relativedelta
from datetime import datetime, timezone, timedelta
from odoo import api, fields, models, _
from odoo.exceptions import UserError, ValidationError

MONTH = [   ('01', 'January'), ('02', 'February'), ('03', 'March'), ('04', 'April'),('05', 'May'),
            ('06', 'June'),('07', 'July'), ('08', 'August'), ('09', 'September'),
            ('10', 'October'), ('11', 'November'), ('12', 'December')]

class Equip3ReportingAssetwizard(models.TransientModel):
    _name = 'equip3.reporting.asset.wizard'
    _description = 'Equip3 Reporting Asset Wizard'

    name = fields.Char('Name')
    json_data = fields.Text('JSON Data')
    filename = fields.Char('Filename')
    datas = fields.Binary('Datas', attachment=False)
    year = fields.Many2one('sh.fiscal.year', string='Year')
    asset_asset_pos = fields.Selection([('Posted','Posted'), ('Non Posted','Non Posted')], string='Asset Entry')
    dept_asset_pos = fields.Selection([('Posted','Posted'), ('Non Posted','Non Posted')], string='Depreciation Entry')

    def action_download(self, workbook, filename=''):
        fd, path = tempfile.mkstemp()
        datas = None
        try:
            workbook.save(path)
            file =  open(path, 'rb').read()
            datas = base64.encodebytes(file)
        finally:
            os.remove(path)

        self.sudo().write({ 'datas': datas, 'filename': filename })
        return {
            'target': 'self',
            'type' : 'ir.actions.act_url',
            'url': f"/reporting-asset/content/{self._name}/{self.id}/datas/{filename}",
        }

    def get_months(self):
        year = self.year.name
        display_month = []
        for month in range(1, 12 + 1):
            _month = "{:02d}".format(month)
            date = datetime.strptime(f'{year}-{_month}-01', '%Y-%m-%d')
            display_month += [{
                'name': date.strftime('%b-%Y'),
                'type': 'calc',
                'field': 'depreciation_expense,' + str(date.strftime('%Y-%m')),
                'alignment': 'right',
                'number_format': True,
            }]
        return display_month

    def _get_asset_value_residual(self):
        query = '''
            SELECT d.asset_id,  sum(d.amount)
            FROM account_asset_depreciation_line AS d
            LEFT JOIN account_move AS am ON am.id = d.move_id
            WHERE (d.move_id IS NULL OR am.state != 'posted')
            GROUP BY d.asset_id
        '''
        self._cr.execute(query)
        return { x[0]: x[1] for x in self._cr.fetchall()}

    def get_datas(self):
        _asset_value_residual = self._get_asset_value_residual()

        where = '''
            WHERE (d.depreciation_date >= '{year}-01-01' AND d.depreciation_date <= '{year}-12-31')
        '''.format(year=self.year.name)

        if self.asset_asset_pos:
            if self.asset_asset_pos == 'Posted':
                where += " AND ( asset.state NOT IN ('draft') AND asset.state IS NOT NULL ) "
            if self.asset_asset_pos == 'Non Posted':
                where += " AND asset.state IN ('draft') "

        query = '''
            SELECT
                asset.id,
                rp.name,
                asset.po_ref,
                asset.asset_sequence,
                asset.location,
                asset.asset_type_1,
                aac.name AS category_id,
                asset.name,
                asset.first_depreciation_manual_date,
                asset.method_number,
                asset.salvage_value,
                asset.value AS gross_value,
                to_char(d.depreciation_date, 'YYYY')
            FROM account_asset_depreciation_line AS d
            INNER JOIN account_asset_asset  AS asset ON asset.id = d.asset_id
            LEFT JOIN res_partner  AS rp ON rp.id = asset.partner_id
            LEFT JOIN account_asset_category  AS aac ON aac.id = asset.category_id
            {where}
            GROUP BY
                asset.id,
                rp.name,
                asset.po_ref,
                asset.asset_sequence,
                asset.location,
                asset.asset_type_1,
                aac.name,
                asset.name,
                asset.first_depreciation_manual_date,
                asset.method_number,
                asset.salvage_value,
                asset.value,
                to_char(d.depreciation_date, 'YYYY')
            ORDER BY
                asset.id,
                to_char(d.depreciation_date, 'YYYY')
        '''.format(where=where)
        self._cr.execute(query)
        results = self._cr.fetchall()

        values = []
        for res in results:
            value = {
                'id': res[0],
                'partner_id': res[1],
                'po_ref': res[2],
                'asset_sequence': res[3],
                'location': res[4],
                'asset_type_1': res[5],
                'category_id': res[6],
                'name': res[7],
                'first_depreciation_manual_date': res[8],
                'method_number': int(res[9]),
                'salvage_value': float(res[10] or 0.0) ,
                'gross_value': float(res[11]), #gross value
            }
            value['asset_value_residual'] = _asset_value_residual.get(value['id'])
            values += [value]
        return values

    def get_value(self, content, assets, context):
        depreciations = context.get('depreciations', {})
        accumulation_previous = context.get('accumulation_previous', {})
        accumulation_additional = context.get('accumulation_additional', {})
        book_year_previous = context.get('book_year_previous', {})
        book_year = context.get('book_year', {})

        value = ''
        if content.get('field') and content.get('type'):
            if content['type'] == 'field':
                if assets.get(content['field']):
                    value = assets[content['field']]

            if content['type'] == 'number':
                if assets.get(content['field']):
                    value = round(assets.get(content['field'], 0.00), 2)
                    if not value:
                        value = ''

            if content['type'] == 'date':
                if assets.get(content['field']):
                    value = assets[content['field']]
                    if value:
                        if type(value) is str:
                            value = datetime.strptime(value, '%Y-%m-%d').strftime('%m/%d/%Y')
                        else:
                            value = value.strftime('%m/%d/%Y')

            if content['type'] == 'calc':
                if content['field'] == 'finish_date':
                    if assets.get('first_depreciation_manual_date') and assets.get('method_number'):
                        value = (assets['first_depreciation_manual_date']
                                + relativedelta(years=assets['method_number'])).strftime('%m/%d/%Y')

                if content['field'] == 'asset_type_1':
                    if assets.get('asset_type_1'):
                        value = assets[content['field']]
                        if value == 'non_building':
                            value = 'Non-Building'

                if content['field'] == 'unit':
                    value = 1

                if content['field'] == 'cost,previous':
                    value = round(assets.get('gross_value', 0.00), 2)
                if content['field'] == 'cost,additional':
                    value = 0
                if content['field'] == 'cost,deduction':
                    value = 0
                if content['field'] == 'cost,ending': # previous + additional - deduction
                    previous = assets.get('gross_value', 0.00)
                    value = round((previous + 0 - 0), 2)

                if content['field'] == 'accu,previous':
                    value = round(accumulation_previous.get(assets['id'], 0.00), 2)
                if content['field'] == 'accu,additional':
                    value = round(accumulation_additional.get(assets['id'], 0.00), 2)
                if content['field'] == 'accu,deduction':
                    value = 0
                if content['field'] == 'accu,ending': # depr + add - deduct
                    previous = round(accumulation_previous.get(assets['id'], 0.00), 2)
                    additional = round(accumulation_additional.get(assets['id'], 0.00), 2)
                    value = round((previous + additional - 0), 2)

                if content['field'] == 'book_year_previous':
                    value = round(book_year_previous.get(assets['id'], 0.00), 2)

                if content['field'] == 'book_year':
                    accu_ending = accumulation_previous.get(assets['id'], 0.0) + accumulation_additional.get(assets['id'], 0.0) - 0
                    gross_value = assets['gross_value']
                    salvage_value = assets['salvage_value']
                    # residual_amount = book_year.get(assets['id'], 0)
                    value = round(gross_value - salvage_value - accu_ending, 2)

                if 'depreciation_expense,' in content['field']:
                    _date = content['field'].replace('depreciation_expense,','')
                    _key = str(assets['id']) + ',' + str(_date)
                    value = round(depreciations.get(_key, 0.00), 2)
                    if not value:
                        value = ''

        return value

    def get_depreciation_datas(self):
        where = '''
            WHERE d.depreciation_date >= '{year}-01-01'
                AND d.depreciation_date <= '{year}-12-31'
        '''.format(year=self.year.name)

        if self.dept_asset_pos:
            if self.dept_asset_pos == 'Posted':
                where += " AND d.move_posted_check = 't' "
            if self.dept_asset_pos == 'Non Posted':
                where += " AND d.move_posted_check = 'f' "

        query = '''
            SELECT
                d.asset_id,
                to_char(d.depreciation_date, 'YYYY-MM') AS date,
                SUM(d.amount)
            FROM account_asset_depreciation_line AS d
            {where}
            GROUP BY
                d.asset_id,
                to_char(d.depreciation_date, 'YYYY-MM')
            ORDER BY d.asset_id, to_char(d.depreciation_date, 'YYYY-MM')
        '''.format(where=where)

        self._cr.execute(query)
        return { str(x[0]) + ',' + str(x[1]): x[2] for x in self._cr.fetchall()}

    def get_accumulation_previous_datas(self):
        # Penyusutan smp periode terpilih
        # Andaikan skrg tahun 2024, maka yang diambil tahun 2023
        where = '''
            WHERE d.depreciation_date < '{year}-01-01'
        '''.format(year=self.year.name)

        if self.dept_asset_pos:
            if self.dept_asset_pos == 'Posted':
                where += " AND d.move_posted_check = 't' "
            if self.dept_asset_pos == 'Non Posted':
                where += " AND d.move_posted_check = 'f' "

        query = '''
            SELECT
                d.asset_id,
                SUM(d.amount)
            FROM account_asset_depreciation_line AS d
            {where}
            GROUP BY
                d.asset_id
            ORDER BY d.asset_id
        '''.format(where=where)

        self._cr.execute(query)
        return { x[0]: x[1] for x in self._cr.fetchall()}

    def get_accumulation_additional_datas(self):
        #Sum all value Depreciation Expense (Montly)
        where = '''
            WHERE d.depreciation_date >= '{year}-01-01'
                AND d.depreciation_date <= '{year}-12-31'
        '''.format(year=self.year.name)

        if self.dept_asset_pos:
            if self.dept_asset_pos == 'Posted':
                where += " AND d.move_posted_check = 't' "
            if self.dept_asset_pos == 'Non Posted':
                where += " AND d.move_posted_check = 'f' "

        query = '''
            SELECT
                d.asset_id,
                SUM(d.amount)
            FROM account_asset_depreciation_line AS d
            {where}
            GROUP BY
                d.asset_id
            ORDER BY d.asset_id
        '''.format(where=where)

        self._cr.execute(query)
        return { x[0]: x[1] for x in self._cr.fetchall()}

    def get_book_year_previous_datas(self):
        # Residual smp periode terpilih
        # Andaikan skrg tahun 2024, maka yang diambil tahun 2023
        query = '''
            SELECT
                d.asset_id,
                SUM(d.remaining_value)
            FROM account_asset_depreciation_line AS d
            WHERE d.depreciation_date < '{year}-01-01'
            GROUP BY
                d.asset_id
            ORDER BY d.asset_id
        '''.format(year=self.year.name)

        self._cr.execute(query)
        return { x[0]: x[1] for x in self._cr.fetchall()}

    def get_book_year_datas(self):
        query = '''
            SELECT
                d.asset_id,
                SUM(d.remaining_value)
            FROM account_asset_depreciation_line AS d
            WHERE d.depreciation_date >= '{year}-01-01'
                AND d.depreciation_date <= '{year}-12-31'
            GROUP BY
                d.asset_id
            ORDER BY d.asset_id
        '''.format(year=self.year.name)

        self._cr.execute(query)
        return { x[0]: x[1] for x in self._cr.fetchall()}


    def action_export_xlsx(self):
        workbook = Workbook()
        sheet = workbook.active

        column_dimensions = [('A',15), ('B',15), ('C',23), ('D',15), ('E',15), ('F',15), ('G',15), ('H',15),
            ('I',15), ('J',15), ('K',15), ('L',15), ('M',15), ('N',15), ('O',15), ('P',15), ('Q',15), ('R',15),
            ('S',15), ('T',15), ('U',15), ('V',15), ('W',15), ('X',15), ('Y',15), ('Z',15), ('AA',15), ('AB',15),
            ('AC',15), ('AD',15), ('AE',15), ('AF',15), ('AG',15), ('AH',15)]
        for dim in column_dimensions:
            sheet.column_dimensions[dim[0]].width = dim[1]

        align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
        align_right = Alignment(horizontal="right", vertical="center", wrap_text=True)
        align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)
        border_thin = Side(border_style="thin", color="000000")

        if not self.year or (self.year and len(self.year.name) != 4 ):
            raise ValidationError(_('Invalid Year.'))

        display_months = self.get_months()
        assets = self.get_datas()

        context = {
            'depreciations': self.get_depreciation_datas(),
            'accumulation_previous': self.get_accumulation_previous_datas(),
            'accumulation_additional': self.get_accumulation_additional_datas(),
            'book_year_previous': self.get_book_year_previous_datas(),
            'book_year': self.get_book_year_datas(),
        }

        row = 1
        column = 1

        row += 1
        cell1 = sheet.cell(row=row, column=column)
        cell1.value = 'Company Active'
        cell1.font = Font(bold=True)
        cell1.alignment = align_left
        sheet.merge_cells(start_row=row, start_column=column, end_row=row, end_column=column + len(column_dimensions) - 1)

        row += 1
        cell1 = sheet.cell(row=row, column=column)
        cell1.value = 'Asset Report'
        cell1.font = Font(bold=True)
        cell1.alignment = align_left
        sheet.merge_cells(start_row=row, start_column=column, end_row=row, end_column=column + len(column_dimensions) - 1)

        column_content = [
            {
                'name': 'Vendor',
                'type': 'field',
                'field': 'partner_id',
                'alignment': 'left',
            },
            {
                'name': 'SPK/PO',
                'type': 'field',
                'field': 'po_ref',
            },
            {
                'name': 'Asset Code',
                'type': 'field',
                'field': 'asset_sequence',
            },
            {
                'name': 'Location',
                'type': 'field',
                'field': 'location',
            },
            {
                'name': 'Asset Type',
                'type': 'calc',
                'field': 'asset_type_1',
            },
            {
                'name': 'Asset classification',
                'type': 'field',
                'field': 'category_id',
            },
            {
                'name': 'Unit',
                'type': 'calc',
                'field': 'unit',  # by default 1
            },
            {
                'name': 'Description',
                'type': 'field',
                'field': 'name',
                'alignment': 'left',
            },
            {
                'name': 'Life Times\n(Years)',
                'type': 'field',
                'field': 'method_number', # Number of Depreciations
            },
            {
                'name': 'Date',
                'childs': [
                    {
                        'name': 'Start Date',
                        'type': 'date',
                        'field': 'first_depreciation_manual_date',
                    },
                    {
                        'name': 'Finish Date',
                        'type': 'calc',
                        'field': 'finish_date',
                    },
                ],
            },
            {
                'name': 'COST',
                'childs': [
                    {
                        'name': 'Previous',
                        'type': 'calc',
                        'field': 'cost,previous',
                        'alignment': 'right',
                        'number_format': True,
                    },
                    {
                        'name': 'Additional',
                        'type': 'calc',
                        'field': 'cost,additional', # by default 0
                        'alignment': 'right',
                        'number_format': True,
                    },
                    {
                        'name': 'Deduction',
                        'type': 'calc',
                        'field': 'cost,deduction', # by default 0
                        'alignment': 'right',
                        'number_format': True,
                    },
                    {
                        'name': 'Ending',
                        'type': 'calc',
                        'field': 'cost,ending', # previous + additional - deduction
                        'alignment': 'right',
                        'number_format': True,
                    },
                ],
            },
            {
                'name': 'Accumulation',
                'childs': [
                    {
                        'name': 'Previous',
                        'type': 'calc',
                        'field': 'accu,previous',
                        'alignment': 'right',
                        'number_format': True,
                    },
                    {
                        'name': 'Additional',
                        'type': 'calc',
                        'field': 'accu,additional',
                        'alignment': 'right',
                        'number_format': True,
                    },
                    {
                        'name': 'Deduction',
                        'type': 'calc',
                        'field': 'accu,deduction',
                        'alignment': 'right',
                        'number_format': True,
                    },
                    {
                        'name': 'Ending',
                        'type': 'calc',
                        'field': 'accu,ending',
                        'alignment': 'right',
                        'number_format': True,
                        # previous + additional - deduction
                    },
                ],
            },
            {
                'name': 'Book Value\n(Previous Year)',
                'type': 'calc',
                'field': 'book_year_previous',
                'alignment': 'right',
                'number_format': True,
            },
            {
                'name': 'Depreciation Expense (Montly)',
                'alignment': 'right',
                'number_format': True,
                'childs': display_months,
            },
            {
                'name': 'BOOK VALUE\n(YTD)',
                'type': 'calc',
                'field': 'book_year',
                'alignment': 'right',
                'number_format': True,
            },
            {
                'name': 'SALVAGE\nVALUE',
                'type': 'field',
                'field': 'salvage_value',
                'alignment': 'right',
                'number_format': True,
            },
        ]

        _column = 1
        _row = row + 3
        for index, cc in enumerate(column_content):
            cell1 = sheet.cell(row=_row, column=_column)
            cell1.value = cc.get('name')
            cell1.alignment = align_center
            cell1.border = Border(top=border_thin, left=border_thin, right=border_thin, bottom=border_thin)
            childs = cc.get('childs')
            if not childs:
                sheet.merge_cells(start_row=_row, start_column=_column, end_row=_row + 1, end_column=_column)
                _column += 1

            if childs:
                sheet.merge_cells(start_row=_row, start_column=_column, end_row=_row, end_column=_column + len(childs) - 1)
                for c in childs:
                    cell2 = sheet.cell(row=_row + 1, column=_column)
                    cell2.value = c.get('name')
                    cell2.alignment = align_center
                    cell2.border = Border(top=border_thin, left=border_thin, right=border_thin, bottom=border_thin)
                    _column += 1

        _row = _row + 2
        for asset in assets:
            _column = 1
            for index, cc in enumerate(column_content):
                cell1 = sheet.cell(row=_row, column=_column)
                childs = cc.get('childs')
                if not childs:
                    cell1.value = self.get_value(cc, asset, context)
                    cell1.border = Border(top=border_thin, left=border_thin, right=border_thin, bottom=border_thin)

                    cell1.alignment = align_center
                    if cc.get('alignment') == 'center':
                        cell1.alignment = align_center
                    if cc.get('alignment') == 'left':
                        cell1.alignment = align_left
                    if cc.get('alignment') == 'right':
                        cell1.alignment = align_right

                    if cc.get('number_format'):
                        cell1.number_format = '#,##0'

                    _column += 1

                if childs:
                    for c in childs:
                        cell2 = sheet.cell(row=_row, column=_column)
                        cell2.value = self.get_value(c, asset, context)
                        cell2.border = Border(top=border_thin, left=border_thin, right=border_thin, bottom=border_thin)

                        cell2.alignment = align_center
                        if c.get('alignment') == 'center':
                            cell2.alignment = align_center
                        if c.get('alignment') == 'left':
                            cell2.alignment = align_left
                        if c.get('alignment') == 'right':
                            cell2.alignment = align_right

                        if c.get('number_format'):
                            cell2.number_format = '#,##0'

                        _column += 1

            _row += 1

        return self.action_download(workbook, filename='Reporting Asset.xlsx')
